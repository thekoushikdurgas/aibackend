"""
Export Handler - Formats and exports analysis data to various formats
"""

import json
import io
import logging
from typing import Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

logger = logging.getLogger(__name__)


class ExportHandler:
    """
    Handles export of analysis data to:
    - JSON
    - CSV (multiple files)
    - Excel (multi-sheet workbook)
    - Markdown
    """

    def __init__(self, analysis_data: Dict[str, Any]):
        """
        Initialize with analysis data.

        Args:
            analysis_data: Complete analysis results
        """
        self.data = analysis_data

    def export_json(self) -> str:
        """
        Export to JSON format.

        Returns:
            JSON string
        """
        return json.dumps(self.data, indent=2, ensure_ascii=False)

    def export_csv(self) -> Dict[str, str]:
        """
        Export to multiple CSV files.
        Similar to Udemy scraper's export_csv method.

        Returns:
            Dictionary mapping filename to CSV content
        """
        csv_files = {}

        # Page info CSV
        if "page_info" in self.data:
            page_info = self.data["page_info"]
            df = pd.DataFrame([page_info])
            csv_files["page_info.csv"] = df.to_csv(index=False)

        # Tables CSV
        if (
            "extracted_content" in self.data
            and "tables" in self.data["extracted_content"]
        ):
            tables = self.data["extracted_content"]["tables"]
            if tables:
                # Flatten all tables into one CSV
                all_rows = []
                for table_idx, table in enumerate(tables):
                    if table.get("headers"):
                        # Add table identifier row
                        all_rows.append(
                            [
                                f"Table {table_idx + 1}"
                                + (
                                    f": {table.get('caption', '')}"
                                    if table.get("caption")
                                    else ""
                                )
                            ]
                            + [""] * (len(table["headers"]) - 1)
                        )
                        all_rows.append(table["headers"])
                        all_rows.extend(table["rows"])
                        all_rows.append([])  # Empty row between tables

                if all_rows:
                    # Find max columns
                    max_cols = max(len(row) for row in all_rows) if all_rows else 0
                    # Pad rows to same length
                    padded_rows = [
                        row + [""] * (max_cols - len(row)) for row in all_rows
                    ]
                    df = pd.DataFrame(padded_rows)
                    csv_files["tables.csv"] = df.to_csv(index=False, header=False)

        # Entities CSV
        if (
            "extracted_content" in self.data
            and "entities" in self.data["extracted_content"]
        ):
            entities = self.data["extracted_content"]["entities"]
            if entities:
                # Flatten entities
                entity_rows = []
                for entity_type, entity_list in entities.items():
                    if isinstance(entity_list, list):
                        for entity in entity_list:
                            if isinstance(entity, dict):
                                entity_rows.append({"type": entity_type, **entity})
                            else:
                                entity_rows.append(
                                    {"type": entity_type, "value": str(entity)}
                                )

                if entity_rows:
                    df = pd.DataFrame(entity_rows)
                    csv_files["entities.csv"] = df.to_csv(index=False)

        # Contact info CSV
        if (
            "extracted_content" in self.data
            and "contact_info" in self.data["extracted_content"]
        ):
            contact_info = self.data["extracted_content"]["contact_info"]
            if contact_info:
                contact_rows = []
                for info_type, info_list in contact_info.items():
                    for info in info_list:
                        contact_rows.append({"type": info_type, "value": info})

                if contact_rows:
                    df = pd.DataFrame(contact_rows)
                    csv_files["contact_info.csv"] = df.to_csv(index=False)

        # Links CSV
        if (
            "extracted_content" in self.data
            and "links" in self.data["extracted_content"]
        ):
            links = self.data["extracted_content"]["links"]
            if links:
                df = pd.DataFrame(links)
                csv_files["links.csv"] = df.to_csv(index=False)

        return csv_files

    def export_excel(self) -> bytes:
        """
        Export to Excel workbook with multiple sheets.
        Similar to Udemy scraper's export_excel method.

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Sheet 1: Page Info
        if "page_info" in self.data:
            ws_page = wb.create_sheet("Page Info")
            page_info = self.data["page_info"]
            headers = list(page_info.keys())
            values = [page_info.get(h) for h in headers]

            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws_page.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            # Write data
            for col_idx, value in enumerate(values, 1):
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                ws_page.cell(row=2, column=col_idx, value=value or "")

            # Auto-adjust column widths
            for col in ws_page.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except (AttributeError, TypeError, ValueError):
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws_page.column_dimensions[column].width = adjusted_width

        # Sheet 2: Tables
        if (
            "extracted_content" in self.data
            and "tables" in self.data["extracted_content"]
        ):
            tables = self.data["extracted_content"]["tables"]
            if tables:
                ws_tables = wb.create_sheet("Tables")
                row_num = 1

                for table_idx, table in enumerate(tables):
                    # Table header
                    ws_tables.cell(
                        row=row_num, column=1, value=f"Table {table_idx + 1}"
                    )
                    if table.get("caption"):
                        ws_tables.cell(row=row_num, column=2, value=table["caption"])
                    row_num += 1

                    # Table headers
                    if table.get("headers"):
                        for col_idx, header in enumerate(table["headers"], 1):
                            cell = ws_tables.cell(
                                row=row_num, column=col_idx, value=header
                            )
                            cell.font = Font(bold=True)
                            cell.fill = PatternFill(
                                start_color="E0E0E0",
                                end_color="E0E0E0",
                                fill_type="solid",
                            )
                        row_num += 1

                    # Table rows
                    for row in table.get("rows", []):
                        for col_idx, cell_value in enumerate(row, 1):
                            ws_tables.cell(
                                row=row_num, column=col_idx, value=cell_value
                            )
                        row_num += 1

                    row_num += 1  # Empty row between tables

                # Auto-adjust column widths
                for col in ws_tables.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except (AttributeError, TypeError, ValueError):
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws_tables.column_dimensions[column].width = adjusted_width

        # Sheet 3: Entities
        if (
            "extracted_content" in self.data
            and "entities" in self.data["extracted_content"]
        ):
            entities = self.data["extracted_content"]["entities"]
            if entities:
                entity_rows = []
                for entity_type, entity_list in entities.items():
                    if isinstance(entity_list, list):
                        for entity in entity_list:
                            if isinstance(entity, dict):
                                entity_rows.append({"type": entity_type, **entity})
                            else:
                                entity_rows.append(
                                    {"type": entity_type, "value": str(entity)}
                                )

                if entity_rows:
                    ws_entities = wb.create_sheet("Entities")
                    df = pd.DataFrame(entity_rows)

                    # Write headers
                    for col_idx, header in enumerate(df.columns, 1):
                        cell = ws_entities.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                        )

                    # Write data
                    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws_entities.cell(row=row_idx, column=col_idx, value=value)

        # Sheet 4: Contact Info
        if (
            "extracted_content" in self.data
            and "contact_info" in self.data["extracted_content"]
        ):
            contact_info = self.data["extracted_content"]["contact_info"]
            if contact_info:
                contact_rows = []
                for info_type, info_list in contact_info.items():
                    for info in info_list:
                        contact_rows.append({"type": info_type, "value": info})

                if contact_rows:
                    ws_contact = wb.create_sheet("Contact Info")
                    df = pd.DataFrame(contact_rows)

                    # Write headers
                    for col_idx, header in enumerate(df.columns, 1):
                        cell = ws_contact.cell(row=1, column=col_idx, value=header)
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(
                            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                        )

                    # Write data
                    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws_contact.cell(row=row_idx, column=col_idx, value=value)

        # Convert to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.read()

    def export_markdown(self) -> str:
        """
        Export to Markdown format.

        Returns:
            Markdown string
        """
        md_parts = []

        # Page info header
        if "page_info" in self.data:
            page_info = self.data["page_info"]
            md_parts.append(f"# {page_info.get('title', 'Page Analysis')}\n")
            md_parts.append(f"\n**URL:** {page_info.get('url', 'N/A')}\n")
            md_parts.append(f"**Page Type:** {page_info.get('page_type', 'N/A')}\n")
            md_parts.append(f"**Confidence:** {page_info.get('confidence', 0)}%\n")

        # AI Analysis summary
        if "ai_analysis" in self.data:
            ai_analysis = self.data["ai_analysis"]
            md_parts.append("\n## AI Analysis\n")
            if ai_analysis.get("summary"):
                md_parts.append(f"{ai_analysis['summary']}\n")
            if ai_analysis.get("insights"):
                md_parts.append("\n### Key Insights\n")
                for insight in ai_analysis["insights"]:
                    md_parts.append(f"- {insight}\n")

        # Headings structure
        if (
            "extracted_content" in self.data
            and "headings" in self.data["extracted_content"]
        ):
            headings_data = self.data["extracted_content"]["headings"]
            if headings_data.get("headings"):
                md_parts.append("\n## Page Structure\n")
                for heading in headings_data["headings"]:
                    level = heading.get("level", 1)
                    text = heading.get("text", "")
                    md_parts.append(f"{'#' * level} {text}\n")

        # Tables
        if (
            "extracted_content" in self.data
            and "tables" in self.data["extracted_content"]
        ):
            tables = self.data["extracted_content"]["tables"]
            if tables:
                md_parts.append("\n## Tables\n")
                for table_idx, table in enumerate(tables, 1):
                    if table.get("caption"):
                        md_parts.append(
                            f"\n### Table {table_idx}: {table['caption']}\n"
                        )
                    else:
                        md_parts.append(f"\n### Table {table_idx}\n")

                    if table.get("headers"):
                        # Create markdown table
                        md_parts.append("| " + " | ".join(table["headers"]) + " |\n")
                        md_parts.append(
                            "| " + " | ".join(["---"] * len(table["headers"])) + " |\n"
                        )
                        for row in table.get("rows", []):
                            md_parts.append(
                                "| " + " | ".join(str(cell) for cell in row) + " |\n"
                            )

        # Contact Info
        if (
            "extracted_content" in self.data
            and "contact_info" in self.data["extracted_content"]
        ):
            contact_info = self.data["extracted_content"]["contact_info"]
            if any(contact_info.values()):
                md_parts.append("\n## Contact Information\n")
                for info_type, info_list in contact_info.items():
                    if info_list:
                        md_parts.append(f"\n### {info_type.title()}\n")
                        for info in info_list:
                            md_parts.append(f"- {info}\n")

        return "\n".join(md_parts)
